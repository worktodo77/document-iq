"""Loader for LI's internal document index (D-04, §5).

The index is maintained by hand in Excel by LI's document database manager, so
the loader treats the *header row* as untrusted: the audited Project 495 index
carries trailing spaces in two column names (``"Original Sort "``,
``"Source Received "``) and an embedded newline in a third (``"Size\\n(KB)"``).
Matching those literally would produce a loader that works on exactly one file.
Headers are therefore normalized and resolved through alias sets, and the
resolution is reported so an operator can see which column DocIQ believed.

The index is a **hashed run input** (D-04 mitigation (a)): the file's SHA-256
and row count go into :class:`dociq.contracts.MasterIndexSnapshot`, which lands
in ``RunConfig`` and in the processing log.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from dociq.contracts import DocIQError, MasterIndexSnapshot

__all__ = [
    "MasterIndexError",
    "MasterIndexRow",
    "QuarantinedRow",
    "MasterIndex",
    "load_master_index",
    "normalize_header",
    "COLUMN_ALIASES",
    "file_sha256",
]

_HASH_CHUNK = 1 << 20


class MasterIndexError(DocIQError):
    """The supplied index cannot be read as a document index.

    Always carries an actionable message: the operator picked the wrong file or
    the sheet's headers changed, and both are fixable without a code change.
    """


def normalize_header(raw: object) -> str:
    """Collapse a spreadsheet header to a comparable token.

    Whitespace of every kind (including the embedded newline in ``Size\\n(KB)``)
    collapses to a single space, surrounding punctuation is dropped, and case is
    folded. Nothing here is specific to the Project 495 file.
    """
    text = "" if raw is None else str(raw)
    text = re.sub(r"\s+", " ", text).strip().casefold()
    text = text.strip(" .:_-")
    return text


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "original_sort": (
        "original sort",
        "originalsort",
        "sort",
        "sort no",
        "sort number",
        "li file no",
        "li file number",
        "file no",
        "file number",
        "index no",
    ),
    "filename": ("filename", "file name", "name", "document name", "doc name"),
    "ext": ("file extension", "extension", "ext", "file type", "filetype"),
    "filepath": (
        "filepath",
        "file path",
        "path",
        "folder",
        "folder path",
        "directory",
        "location",
        "relative path",
    ),
    "size_kb": ("size (kb)", "size kb", "size(kb)", "size", "file size (kb)", "file size"),
    "date_text": ("date", "document date", "doc date", "file date"),
    "source_received": ("source received", "source", "received from"),
    "date_received": ("date received", "received", "received date"),
    "sha256": ("sha256", "sha-256", "sha 256", "hash", "sha256 hash", "checksum", "md5"),
    "bates_start": (
        "bates start",
        "bates begin",
        "beg bates",
        "begin bates",
        "bates beg",
        "begbates",
        "bates from",
    ),
    "bates_end": ("bates end", "end bates", "endbates", "bates to", "bates thru"),
    "bates_range": ("bates", "bates range", "bates number", "bates numbers"),
}
"""Alias sets, most specific first within each field.

``sha256`` deliberately accepts ``md5`` as an alias: an index that carries any
content digest is more useful as a secondary match key than none, and the
assigner compares whatever the column holds against the same-algorithm digest it
is given. The column's real name is preserved in the resolution report so the
operator is never misled about what was compared.
"""

_REQUIRED = ("original_sort", "filename", "filepath")


@dataclass(frozen=True, slots=True)
class MasterIndexRow:
    """One index row, with every raw cell retained.

    ``raw`` keeps the untouched (header, value) pairs so the durable-artifact
    rule holds: everything the file said is hashable and reproducible, not just
    the fields DocIQ happens to consume today.
    """

    row_number: int
    """1-based position among data rows, independent of ``original_sort``."""

    original_sort: int
    filename: str
    filepath: str
    ext: str | None = None
    size_kb: int | None = None
    date_text: str | None = None
    source_received: str | None = None
    date_received: str | None = None
    sha256: str | None = None
    bates_start: str | None = None
    bates_end: str | None = None
    raw: tuple[tuple[str, str], ...] = ()

    @property
    def li_file_no(self) -> str:
        """The index's own identifier for this row, as a string (§5)."""
        return str(self.original_sort)


QUARANTINE_NO_SORT = "no usable Original Sort value"
QUARANTINE_NEGATIVE_SORT = "negative Original Sort value"
QUARANTINE_NO_FILENAME = "no filename"
QUARANTINE_DUPLICATE_SORT = "duplicate Original Sort value"
"""The complete set of reasons a data row can be held out of the LI number
space. Named rather than spelled inline so the loader's warning, the retained
row and the reconciliation entry cannot drift apart."""


@dataclass(frozen=True, slots=True)
class QuarantinedRow:
    """An index row the loader could not admit to the LI number space.

    D-1: the loader warns about these rows and states they will reconcile as
    index-only. Retaining them is what makes that statement true. They keep
    every raw cell — the operator's job is to repair the *spreadsheet*, and the
    only way to find the offending cell is to be shown what it said.

    A quarantined row deliberately has **no** ``original_sort`` and **no**
    ``li_file_no``: it never entered the identifier space, and rendering an
    invented one would be worse than rendering none.
    """

    row_number: int
    """1-based position among data rows, on the same counter as
    :attr:`MasterIndexRow.row_number`, so the two sets never collide."""

    reason: str
    """One of the ``QUARANTINE_*`` constants."""

    original_sort_text: str
    """Whatever the Original Sort cell actually said, verbatim (possibly '')."""

    filename: str
    filepath: str
    ext: str | None = None
    size_kb: int | None = None
    duplicate_of_row: int | None = None
    """For :data:`QUARANTINE_DUPLICATE_SORT`, the row that claimed the value
    first; ``None`` otherwise."""

    raw: tuple[tuple[str, str], ...] = ()

    @property
    def detail(self) -> str:
        """One operator-facing sentence, safe to put in a spreadsheet cell."""
        text = (
            f"index row {self.row_number} was not assigned an LI File No: "
            f"{self.reason}"
        )
        if self.original_sort_text:
            text += f" (Original Sort cell read {self.original_sort_text!r})"
        if self.duplicate_of_row is not None:
            text += f"; first claimed by index row {self.duplicate_of_row}"
        return text


@dataclass(frozen=True, slots=True)
class MasterIndex:
    """A loaded index plus everything needed to explain how it was read."""

    snapshot: MasterIndexSnapshot
    rows: tuple[MasterIndexRow, ...]
    source_path: str
    sheet_name: str | None
    resolved_columns: tuple[tuple[str, str], ...]
    """``(logical field, header as it appeared in the file)`` — the audit trail
    for the alias resolution."""
    unmapped_headers: tuple[str, ...]
    warnings: tuple[str, ...] = ()
    quarantined: tuple[QuarantinedRow, ...] = ()
    """Rows retained for reconciliation but held out of the LI number space.

    Deliberately a *separate* collection from :attr:`rows`. Every property and
    lookup below reads :attr:`rows` only, so a quarantined row cannot reach
    :attr:`max_original_sort`, :meth:`by_original_sort`, :attr:`has_hashes`,
    :attr:`has_bates`, or the assigner — it has no identifier to lend.
    """

    @property
    def quarantined_count(self) -> int:
        return len(self.quarantined)

    @property
    def has_hashes(self) -> bool:
        return any(r.sha256 for r in self.rows)

    @property
    def has_bates(self) -> bool:
        return any(r.bates_start or r.bates_end for r in self.rows)

    @property
    def max_original_sort(self) -> int:
        return max((r.original_sort for r in self.rows), default=0)

    def by_original_sort(self) -> dict[int, MasterIndexRow]:
        return {r.original_sort: r for r in self.rows}


def file_sha256(path: Path) -> str:
    """Streamed digest — index files are small today, but the same helper is
    used for corpus files and must not load a gigabyte into memory."""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(_HASH_CHUNK), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        # Excel hands back midnight-stamped datetimes for date-only cells;
        # rendering the time would put a spurious "00:00:00" into the log.
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _as_int(value: object) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        try:
            f = float(text)
        except ValueError:
            return None
        return int(f) if f.is_integer() else None


def _resolve_columns(
    headers: Sequence[object],
) -> tuple[dict[str, int], list[tuple[str, str]], list[str]]:
    """Map logical field names onto column positions.

    A field claims the *first* column whose normalized header appears in its
    alias set, scanning aliases in declared order so that a specific alias
    ("size (kb)") wins over a generic one ("size") even when both are present.
    A column is claimed by at most one field.
    """
    normalized = [normalize_header(h) for h in headers]
    taken: set[int] = set()
    mapping: dict[str, int] = {}
    resolved: list[tuple[str, str]] = []
    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for col, norm in enumerate(normalized):
                if col in taken or norm != alias:
                    continue
                mapping[field_name] = col
                taken.add(col)
                resolved.append((field_name, str(headers[col])))
                break
            if field_name in mapping:
                break
    unmapped = [
        str(headers[i]) for i, norm in enumerate(normalized) if i not in taken and norm
    ]
    return mapping, resolved, unmapped


def _read_xlsx(path: Path) -> tuple[str, list[list[object]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - declared dependency
        raise MasterIndexError(
            "Reading an .xlsx index requires openpyxl, which is a declared "
            "dependency of DocIQ; the installation is incomplete."
        ) from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise MasterIndexError(f"{path.name}: workbook contains no sheets.")
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return wb.sheetnames[0], rows
    finally:
        wb.close()


def _read_xls(path: Path) -> tuple[str, list[list[object]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - declared dependency
        raise MasterIndexError(
            "Reading a legacy .xls index requires xlrd, which is a declared "
            "dependency of DocIQ; the installation is incomplete."
        ) from exc
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    return sheet.name, rows


def _read_csv(path: Path) -> tuple[None, list[list[object]]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - cp1252 decodes any byte string
        raise MasterIndexError(f"{path.name}: could not be decoded as text.")
    sample = text[:8192]
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample, ",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text, newline=""), dialect)
    return None, [list(r) for r in reader]


def load_master_index(path: str | Path) -> MasterIndex:
    """Read an index from ``.xlsx``, ``.xls`` or ``.csv``.

    Raises :class:`MasterIndexError` with an operator-actionable message when
    the file cannot be understood. A row that is individually unusable (no
    usable Original Sort, a negative one, no filename, or an Original Sort some
    earlier row already claimed) is held out of the LI number space *with a
    warning*, never silently — and is retained as a :class:`QuarantinedRow` so
    it still surfaces as an "in index, not in folder" entry that an operator
    would otherwise chase for hours.

    Wholly blank spreadsheet rows are the one thing dropped outright: they
    carry no cell an operator could act on, and Excel appends them by the
    thousand.
    """
    p = Path(path)
    if not p.is_file():
        raise MasterIndexError(f"Master index not found: {p}")
    suffix = p.suffix.lower()
    if suffix == ".xlsx" or suffix == ".xlsm":
        sheet_name, grid = _read_xlsx(p)
    elif suffix == ".xls":
        sheet_name, grid = _read_xls(p)
    elif suffix in (".csv", ".txt", ".tsv"):
        sheet_name, grid = _read_csv(p)
    else:
        raise MasterIndexError(
            f"{p.name}: unsupported index format {suffix!r}. "
            "Supply the index as .xlsx, .xls or .csv."
        )

    grid = [row for row in grid if any(_cell_text(c) for c in row)]
    if not grid:
        raise MasterIndexError(f"{p.name}: the index is empty.")

    headers = grid[0]
    mapping, resolved, unmapped = _resolve_columns(headers)
    missing = [f for f in _REQUIRED if f not in mapping]
    if missing:
        seen = ", ".join(repr(str(h)) for h in headers if _cell_text(h))
        raise MasterIndexError(
            f"{p.name}: the index is missing required column(s) {missing}. "
            f"Columns found: {seen}. Expected an 'Original Sort' (or 'LI File No'), "
            "a 'Filename' and a 'Filepath' column."
        )

    header_names = [str(h) if h is not None else "" for h in headers]
    rows: list[MasterIndexRow] = []
    quarantined: list[QuarantinedRow] = []
    warnings: list[str] = []
    seen_sort: dict[int, int] = {}

    def cell(row: Sequence[object], name: str) -> object:
        col = mapping.get(name)
        if col is None or col >= len(row):
            return None
        return row[col]

    def raw_cells(row: Sequence[object]) -> tuple[tuple[str, str], ...]:
        return tuple(
            (header_names[c] if c < len(header_names) else f"column {c + 1}",
             _cell_text(v) or "")
            for c, v in enumerate(row)
        )

    def quarantine(
        row: Sequence[object],
        i: int,
        reason: str,
        *,
        duplicate_of: int | None = None,
    ) -> None:
        """Hold a row out of the LI number space without losing it."""
        quarantined.append(
            QuarantinedRow(
                row_number=i,
                reason=reason,
                original_sort_text=_cell_text(cell(row, "original_sort")) or "",
                filename=_cell_text(cell(row, "filename")) or "",
                filepath=_cell_text(cell(row, "filepath")) or "",
                ext=_cell_text(cell(row, "ext")),
                size_kb=_as_int(cell(row, "size_kb")),
                duplicate_of_row=duplicate_of,
                raw=raw_cells(row),
            )
        )
        detail = f" (first seen at row {duplicate_of})" if duplicate_of else ""
        warnings.append(
            f"index row {i}: {reason}{detail} — the row is kept out of the LI "
            "number space and is reported as index-only in reconciliation"
        )

    for i, row in enumerate(grid[1:], start=1):
        raw_sort = _as_int(cell(row, "original_sort"))
        sort_value = raw_sort
        negative = raw_sort is not None and raw_sort < 0
        if negative:
            # A negative Original Sort cannot become a Doc ID (DocId.base is
            # non-negative by construction, D-04) and a hand-typed Excel cell
            # can easily carry one (a stray "-", a fat-fingered minus). Keep it
            # out of the number space rather than let it reach
            # dociq.docid.ids and crash the whole run — one bad cell must not
            # take down every other document's assignment (Principle 1).
            sort_value = None
        filename = _cell_text(cell(row, "filename"))
        if sort_value is None:
            quarantine(
                row,
                i,
                QUARANTINE_NEGATIVE_SORT if negative else QUARANTINE_NO_SORT,
            )
            continue
        if filename is None:
            quarantine(row, i, QUARANTINE_NO_FILENAME)
            continue
        prior = seen_sort.get(sort_value)
        if prior is not None:
            quarantine(row, i, QUARANTINE_DUPLICATE_SORT, duplicate_of=prior)
            continue
        seen_sort[sort_value] = i
        bates_start = _cell_text(cell(row, "bates_start"))
        bates_end = _cell_text(cell(row, "bates_end"))
        if bates_start is None and bates_end is None:
            combined = _cell_text(cell(row, "bates_range"))
            if combined:
                # The word separators need surrounding whitespace: an
                # unanchored "to" would split "PHOTO 1" into "PHO" and " 1".
                parts = re.split(
                    r"\s*[-–—]\s*|\s+(?:through|thru|to)\s+",
                    combined,
                    maxsplit=1,
                    flags=re.IGNORECASE,
                )
                bates_start = parts[0].strip() or None
                bates_end = (parts[1].strip() if len(parts) > 1 else parts[0].strip()) or None
        raw = raw_cells(row)
        rows.append(
            MasterIndexRow(
                row_number=i,
                original_sort=sort_value,
                filename=filename,
                filepath=_cell_text(cell(row, "filepath")) or "",
                ext=_cell_text(cell(row, "ext")),
                size_kb=_as_int(cell(row, "size_kb")),
                date_text=_cell_text(cell(row, "date_text")),
                source_received=_cell_text(cell(row, "source_received")),
                date_received=_cell_text(cell(row, "date_received")),
                sha256=(_cell_text(cell(row, "sha256")) or "").lower() or None,
                bates_start=bates_start,
                bates_end=bates_end,
                raw=raw,
            )
        )

    if not rows:
        raise MasterIndexError(
            f"{p.name}: no usable rows. Every row lacked an Original Sort value "
            "or a filename, carried a negative Original Sort, or repeated an "
            "Original Sort an earlier row had already used."
        )

    # ``row_count`` deliberately keeps the meaning it has always had: the
    # number of rows admitted to the LI number space, i.e. ``len(rows)``.
    # Widening it to include quarantined rows would silently change a figure
    # that is already printed in the processing log and in the workbook's
    # "Master index rows" line, and would make it disagree with
    # ``max_original_sort`` and the assigner's match denominator. The
    # quarantined population is reported separately, through
    # ``MasterIndex.quarantined_count``, the loader warnings, and the
    # reconciliation's index-only rows.
    snapshot = MasterIndexSnapshot(
        filename=p.name, sha256=file_sha256(p), row_count=len(rows)
    )
    if quarantined:
        warnings.append(
            f"{len(quarantined)} of {len(rows) + len(quarantined)} index data "
            f"row(s) could not be given an LI File No; 'Master index rows' "
            f"below counts the {len(rows)} usable row(s) only"
        )
    return MasterIndex(
        snapshot=snapshot,
        rows=tuple(rows),
        source_path=str(p),
        sheet_name=sheet_name,
        resolved_columns=tuple(resolved),
        unmapped_headers=tuple(unmapped),
        warnings=tuple(warnings),
        quarantined=tuple(quarantined),
    )

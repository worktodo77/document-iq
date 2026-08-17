"""``document_index.csv`` and ``document_index.xlsx`` — the §5 deliverable.

§5 calls the index "a first-class deliverable, not a by-product", so the two
formats have different jobs and different guarantees:

* **CSV** is machine-readable and *is* inside the byte-identical claim. Written
  by hand rather than through :mod:`csv`'s writer so the line terminator cannot
  be platform-dependent.
* **XLSX** is the formatted, LI-styled copy for humans, and is explicitly
  **outside** the byte-identical claim: the format embeds a creation timestamp
  in its package metadata, so two runs cannot produce identical bytes. The
  freeze document states that split, and so does the hash manifest.

The reconciliation report (§5) is a separate tab of the workbook, and a separate
CSV beside the main one so the machine-readable half is not lost.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence

from dociq.contracts import (
    Disposition,
    DocumentRecord,
    ProcessingStatus,
    document_sort_key,
)
from dociq.docid.reconcile import ReconciliationReport
from dociq.emit.paths import OutputLayout, write_text_deterministic
from dociq.identify.bates import BatesRange
from dociq.sections.apply import SectionDropEntry

__all__ = [
    "INDEX_COLUMNS",
    "RECONCILIATION_COLUMNS",
    "IndexRow",
    "build_index_rows",
    "render_index_csv",
    "write_index_csv",
    "write_reconciliation_csv",
    "write_index_xlsx",
    "LI_NAVY",
    "LI_BLUE",
]

LI_NAVY = "0E4D80"
LI_BLUE = "2E9FD4"
"""D-07's ruled palette. Hex without the leading '#' because that is openpyxl's
form."""

INDEX_COLUMNS = (
    "Doc ID",
    "LI File No",
    "Filename",
    "Relative path",
    "Format",
    "Date (detected)",
    "Document type",
    "Page count",
    "Bates start",
    "Bates end",
    "SHA-256",
    "Parent doc",
    "Processing status",
    "Sections recognized",
    "Recognition tier",
    "Sections dropped",
    "Omission approved by",
)
"""§5's field list, in §5's order. "Relative path" is the one addition: without
it the reconciliation tab's folder-side entries cannot be located on disk, and
every consumer of the index otherwise has to re-derive it.

**Three more were added when the taxonomy landed**, and each answers a question
"Sections dropped" raises and cannot settle. A count of omissions is unreadable
without the recognition it came out of — a document showing 0 dropped is either
a document nothing was recognized in or a document an expert chose to keep whole,
and those are different facts. "Recognition tier" carries §5.4/A-18 to the
artifact a reviewer actually opens: the document's own outline and a page-class
rule are not equally strong evidence, and an index that renders them identically
is what §3 calls the quiet lie in this feature. "Omission approved by" carries
D-34 the same way — the name is a person who acted, and an empty cell means
nothing was omitted rather than that nobody is answerable."""

RECONCILIATION_COLUMNS = (
    "Category",
    "Doc ID",
    "LI File No",
    "Filename",
    "Path",
    "Match method",
    "Field",
    "Folder value",
    "Index value",
    "Detail",
)


@dataclass(frozen=True, slots=True)
class IndexRow:
    """One row of the index deliverable, already stringified.

    Stringified at construction so the CSV and the workbook cannot disagree
    about how a value renders — a number formatted one way in one output and
    another way in the other is the kind of discrepancy that costs an afternoon.
    """

    values: tuple[str, ...]

    def __post_init__(self) -> None:
        if len(self.values) != len(INDEX_COLUMNS):
            raise ValueError(
                f"index row has {len(self.values)} values, expected "
                f"{len(INDEX_COLUMNS)}"
            )


def _status_label(status: ProcessingStatus) -> str:
    return {
        ProcessingStatus.FULL: "Full",
        ProcessingStatus.PARTIAL_OCR_FLAGGED: "Partial-OCR-flagged",
        ProcessingStatus.UNSUPPORTED: "Unsupported",
        ProcessingStatus.FAILED: "Failed",
    }[status]


def build_index_rows(
    documents: Sequence[DocumentRecord],
    *,
    bates_ranges: Mapping[tuple[str, str, int], BatesRange] | None = None,
    drops: Sequence[SectionDropEntry] = (),
) -> tuple[IndexRow, ...]:
    """Build the index rows in canonical document order.

    ``drops`` supplies the "Omission approved by" column and has to be passed
    in: the approver is deliberately NOT on :class:`PageRecord`. A page records
    which rule dropped it (``drop_rule``, ``template_id:family_id``); the person
    who approved the omission lives on the :class:`ApprovedOmission` in the
    matter folder, because D-34 makes an approval a matter-scoped act by a named
    human rather than a property of a page. Deriving a name from ``drop_rule``
    would print a template id under a column headed "approved by", which is the
    class of misdescription D-35 exists to remove.
    """
    ranges = bates_ranges or {}
    approvers_by_doc: dict[str, set[str]] = {}
    for entry in drops:
        approvers_by_doc.setdefault(entry.doc_id, set()).add(entry.approved_by)
    rows: list[IndexRow] = []
    for doc in sorted(documents, key=document_sort_key):
        rng = ranges.get(document_sort_key(doc))
        dropped_sections = {
            p.section or p.drop_rule or ""
            for p in doc.pages
            if p.disposition is Disposition.DROP
        }
        recognized = {p.section for p in doc.pages if p.section is not None}
        # Tiers are rendered strongest-first and de-duplicated, not counted: a
        # document is commonly placed by its outline in one part and by a
        # page-class rule in another, and a reader needs to know BOTH were used
        # rather than which was used more.
        tiers = sorted(
            {p.section_tier.value for p in doc.pages if p.section_tier is not None}
        )
        approvers = sorted(approvers_by_doc.get(doc.doc_id, ()))
        rows.append(
            IndexRow(
                (
                    doc.doc_id,
                    doc.li_file_no or "",
                    doc.filename,
                    doc.rel_path,
                    doc.ext.lstrip("."),
                    "; ".join(doc.detected_dates),
                    doc.doc_type or "",
                    str(doc.pages_in),
                    (rng.start if rng else None) or "",
                    (rng.end if rng else None) or "",
                    doc.sha256,
                    doc.parent_doc_id or "",
                    _status_label(doc.status),
                    str(len(recognized)),
                    "; ".join(tiers),
                    str(len(dropped_sections)),
                    "; ".join(approvers),
                )
            )
        )
    return tuple(rows)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _csv_field(value: str) -> str:
    """RFC 4180 quoting, applied by hand.

    :mod:`csv` would do this too, but its writer takes its line terminator from
    the dialect and its file handle from the caller, which is two more ways for
    a CRLF to reach a file that is inside the byte-identical claim.
    """
    text = value.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    if any(c in text for c in (',', '"', "\n")):
        return '"' + text.replace('"', '""') + '"'
    return text


def _csv_line(values: Sequence[str]) -> str:
    return ",".join(_csv_field(v) for v in values) + "\n"


def render_index_csv(rows: Sequence[IndexRow]) -> str:
    out = [_csv_line(INDEX_COLUMNS)]
    out.extend(_csv_line(r.values) for r in rows)
    return "".join(out)


def write_index_csv(rows: Sequence[IndexRow], layout: OutputLayout) -> Path:
    return write_text_deterministic(layout.index_csv, render_index_csv(rows))


def _reconciliation_rows(report: ReconciliationReport) -> tuple[tuple[str, ...], ...]:
    """Flatten the §5 report into rows.

    A matched pair with three discrepancies becomes three rows, one per field,
    because a spreadsheet reader filters by column and a packed "3 problems"
    cell cannot be filtered.
    """
    out: list[tuple[str, ...]] = []
    for pair in report.matched:
        if not pair.discrepancies:
            out.append(
                (
                    "In both",
                    pair.doc_id,
                    pair.li_file_no,
                    "",
                    pair.rel_path,
                    pair.match_method,
                    "",
                    "",
                    "",
                    "",
                )
            )
            continue
        for d in pair.discrepancies:
            out.append(
                (
                    "In both — discrepancy",
                    pair.doc_id,
                    pair.li_file_no,
                    "",
                    pair.rel_path,
                    pair.match_method,
                    d.field,
                    d.folder_value,
                    d.index_value,
                    d.detail or "",
                )
            )
    for entry in report.folder_only:
        out.append(
            (
                "In folder, not in index",
                entry.doc_id,
                "",
                entry.filename,
                entry.rel_path,
                "",
                "",
                f"{entry.size_bytes} bytes",
                "",
                entry.reason,
            )
        )
    for entry in report.index_only:
        # A quarantined row gets its own category and an empty "LI File No"
        # cell. Reusing the ordinary category would tell the operator the
        # client failed to send a numbered document; the truth is that the
        # index cell itself is unusable, and the fix is in the spreadsheet.
        out.append(
            (
                "In index, unusable row" if entry.quarantined
                else "In index, not in folder",
                "",
                entry.li_file_no,
                entry.filename or f"(index row {entry.index_row_number})",
                entry.filepath,
                "",
                "",
                "",
                f"{entry.size_kb} KB" if entry.size_kb is not None else "",
                entry.reason or "",
            )
        )
    return tuple(out)


def write_reconciliation_csv(
    report: ReconciliationReport, layout: OutputLayout
) -> Path:
    rows = _reconciliation_rows(report)
    text = _csv_line(RECONCILIATION_COLUMNS) + "".join(_csv_line(r) for r in rows)
    return write_text_deterministic(layout.root / "reconciliation.csv", text)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _style_header(ws, columns: Sequence[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=LI_NAVY)
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, columns: Sequence[str], rows: Sequence[Sequence[str]]) -> None:
    from openpyxl.utils import get_column_letter

    for col, name in enumerate(columns, start=1):
        widest = max(
            [len(name)] + [len(r[col - 1]) for r in rows if col - 1 < len(r)] or [len(name)]
        )
        # Capped, and the cap is stated: a 64-character SHA-256 column would
        # otherwise push every other column off the screen.
        ws.column_dimensions[get_column_letter(col)].width = min(max(widest + 2, 10), 48)


def write_index_xlsx(
    rows: Sequence[IndexRow],
    layout: OutputLayout,
    report: ReconciliationReport | None = None,
    *,
    matter_name: str = "",
) -> Path:
    """Write the LI-styled workbook.

    Outside the byte-identical claim by construction — the xlsx container
    records a creation time. That is stated here, in the freeze document, and in
    the hash manifest, rather than left for someone to discover by diffing two
    runs.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - declared dependency
        raise RuntimeError(
            "Writing document_index.xlsx requires openpyxl, a declared "
            "dependency of DocIQ; the installation is incomplete."
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Document Index"
    _style_header(ws, INDEX_COLUMNS)
    body = [r.values for r in rows]
    for r, values in enumerate(body, start=2):
        for c, value in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=value)
    _autosize(ws, INDEX_COLUMNS, body)
    if body:
        from openpyxl.utils import get_column_letter

        last_col = get_column_letter(len(INDEX_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(body) + 1}"

    if report is not None:
        rec = wb.create_sheet("Reconciliation")
        _style_header(rec, RECONCILIATION_COLUMNS)
        rec_rows = _reconciliation_rows(report)
        for r, values in enumerate(rec_rows, start=2):
            for c, value in enumerate(values, start=1):
                rec.cell(row=r, column=c, value=value)
        _autosize(rec, RECONCILIATION_COLUMNS, rec_rows)

        notes = wb.create_sheet("Reconciliation notes")
        _style_header(notes, ("Item", "Value"))
        summary: list[tuple[str, str]] = [
            ("Matter", matter_name),
            ("Master index", report.snapshot.filename if report.snapshot else ""),
            ("Master index SHA-256", report.snapshot.sha256 if report.snapshot else ""),
            ("Master index rows", str(report.snapshot.row_count) if report.snapshot else ""),
            ("Folder root aligned at", report.root_prefix or "(index root)"),
        ]
        summary.extend((k.replace("_", " ").title(), str(v)) for k, v in report.totals.items())
        summary.extend(("Warning", w) for w in report.warnings)
        for r, (k, v) in enumerate(summary, start=2):
            notes.cell(row=r, column=1, value=k)
            notes.cell(row=r, column=2, value=v)
        _autosize(notes, ("Item", "Value"), [(k, v) for k, v in summary])

    layout.root.mkdir(parents=True, exist_ok=True)
    wb.save(layout.index_xlsx)
    return layout.index_xlsx
